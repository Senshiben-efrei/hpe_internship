
# import libraries
import pandas as pd
import numpy as np
import streamlit as st
from discord_webhook import DiscordWebhook
# set to wide view
st.set_page_config(layout="wide")

# create a streamlit app
st.title('Order Suggestion App')


# create 4 columns
col1, col2, col3, col4 = st.columns(4)


# load excel data
catalog_file = col1.file_uploader('catalog_file', type='xlsx', key='catalog_file')
if catalog_file is not None:
    catalog = pd.read_excel(catalog_file)
inventory_file = col2.file_uploader('inventory_file', type='xlsx', key='inventory_file')
if inventory_file is not None:
    inventory = pd.read_excel(inventory_file)
sellout_file = col3.file_uploader('sellout_file', type='xlsx', key='sellout_file')
if sellout_file is not None:
    sellout = pd.read_excel(sellout_file)
    
# chose a file
backlog_backups_file = col4.file_uploader('backlog', type='csv', key='backlog_backups_file')
# if the user chose a file
if backlog_backups_file is not None:
    backlog_backups = pd.read_csv(backlog_backups_file)
    # drop Serial Number and Box ID rows
    backlog_backups = backlog_backups.drop(['Serial Number', 'Box ID'], axis=1)
    # drop duplicates rows
    backlog_backups = backlog_backups.drop_duplicates()

    backlog = backlog_backups 

# if all the files are uploaded
if backlog_backups_file is not None and catalog_file is not None and inventory_file is not None and sellout_file is not None:
    # send report 
    webhook = DiscordWebhook(url='https://discord.com/api/webhooks/1064563281798381708/nBbWnB5oQ-gYxU2m3ZerT_nRP5QjlHLd3thLLM2Bg6S3N93ZPnUnnOp4ww5T3kJYdEFM', content='initializing')
    response = webhook.execute()
    #cleaning
    # change date format to datetime
    sellout['Date'] = pd.to_datetime(sellout['Date'])
    # if the part number contains ' 0D1' or ' OD1' then replace it with ''
    sellout['Part number'] = sellout['Part number'].str.replace(' 0D1', '')
    sellout['Part number'] = sellout['Part number'].str.replace(' OD1', '')
    # in the distributor column, uppercase all the letters
    sellout['Distributor'] = sellout['Distributor'].str.upper()
    # remove the leading and trailing spaces from the distributor column
    sellout['Distributor'] = sellout['Distributor'].str.strip()
    # remove the leading and trailing spaces from the part number column
    sellout['Part number'] = sellout['Part number'].str.strip()

    # Set the index of the dataframe to the 'Date' column
    sellout.set_index('Date', inplace=True)

    # Group the data by 'Part number' and 'Distributor'
    grouped_data = sellout.groupby(['Part number', 'Distributor'])

    # Resample the data by week
    resampled_data = grouped_data.resample('W')

    # Aggregate the data by summing the 'Sold Qty' column
    aggregated_data = resampled_data['Sold Qty'].sum()

    # look for the last 13 weeks from today and average the data for each part number and distributor
    # Get todays date and subtract 13 weeks
    last_13_weeks = pd.to_datetime('today') - pd.DateOffset(weeks=13)

    # remove the data whose date is less than 13 weeks from today
    filtered_data = aggregated_data[aggregated_data.index.get_level_values('Date') > last_13_weeks]

    # average the data for each part number and distributor
    average_data = filtered_data.groupby(['Part number', 'Distributor']).sum()

    # Reset the index of the dataframe
    average_data = average_data.reset_index()

    # store the average data in a new dataframe with the columns 'Part number', 'Distributor', and 'Average'
    average_data = average_data[['Part number', 'Distributor', 'Sold Qty']]
    # rename the 'Part number' column to 'PartID'
    average_data.rename(columns={'Part number': 'PartID'}, inplace=True)



    # in the catalog dataframe drop the Row Labels column
    catalog.drop('Row Labels', axis=1, inplace=True)


    # drop duplicate rows
    catalog.drop_duplicates(inplace=True)


    # process the backlog data
    backlog['Actual Delivery Date (Item)'] = pd.to_datetime(backlog['Actual Delivery Date (Item)'])
    backlog['HPE Received Date'] = pd.to_datetime(backlog['HPE Received Date'])

    # calculate the the time between the HPE Received Date and the Actual Delivery Date (Item)
    backlog['time_to_ship'] = backlog['Actual Delivery Date (Item)'] - backlog['HPE Received Date'] 

    # convert the time to days
    backlog['time_to_ship'] = backlog['time_to_ship'].dt.days/5

    # drop rows where time_to_ship is inf or nan
    backlog = backlog[backlog['time_to_ship'] != np.inf]
    backlog = backlog[backlog['time_to_ship'].notna()]

    # drop rows where Order Type is not 'HPE Non Stkng Order'
    backlog = backlog[backlog['Order Type'] == 'HPE Stocking Order']

    # drop all comlumns except for 'Part number', 'shiped_to', 'time_to_ship', 'Material Description'
    backlog = backlog[['Product Number', 'time_to_ship', 'Product Description']]

    # calculate the average time to ship for each part number keeping the product description
    backlog = backlog.groupby(['Product Number', 'Product Description']).median()

    # reset the index keep product description
    backlog = backlog.reset_index()

    # drop all rows where product number is not in the catalog
    backlog = backlog[backlog['Product Number'].isin(catalog['PartID'])]

    # rename the 'Product Number' column to 'PartID'
    backlog.rename(columns={'Product Number': 'PartID'}, inplace=True)
    backlog_backups.rename(columns={'Product Number': 'PartID'}, inplace=True)

    # in backlog_backups, change Item Status to 'Delivered' if the Actual Delivery Date (Item) is not null
    backlog_backups['Item Status'] = np.where(backlog_backups['Actual Delivery Date (Item)'].notna(), 'Delivered', backlog_backups['Item Status'])



    # in inventory dataframe keep this columns  Week	Month	Date	Distributor	BU	Part number	Inventory( Units)
    inventory = inventory[['Week', 'Month', 'Date', 'Distributor', 'BU', 'Part number', 'Inventory( Units)']]

    # if the part number contains ' 0D1' or ' OD1' then replace it with ''
    inventory['Part number'] = inventory['Part number'].str.replace(' 0D1', '')
    inventory['Part number'] = inventory['Part number'].str.replace(' OD1', '')
    # in the distributor column, uppercase all the letters
    inventory['Distributor'] = inventory['Distributor'].str.upper()
    # remove the leading and trailing spaces from the distributor column
    inventory['Distributor'] = inventory['Distributor'].str.strip()
    # remove the leading and trailing spaces from the part number column
    inventory['Part number'] = inventory['Part number'].str.strip()
    

    # remove the first letter in week column 
    inventory['Week'] = inventory['Week'].str[1:]

    # convert the week column to integer
    inventory['Week'] = inventory['Week'].astype(int)

    # look for the latest Week in the inventory dataframe
    latest_week = inventory['Week'].max()

    # filter the inventory dataframe to keep only the latest week
    inventory = inventory[inventory['Week'] == latest_week]

    # limit the inventory dataframe to catalog part numbers
    inventory = inventory[inventory['Part number'].isin(catalog['PartID'])]

    # rename the column part number to PartID
    inventory.rename(columns={'Part number': 'PartID'}, inplace=True)

    # drop Week	Month	Date		BU
    inventory.drop(['Week', 'Month', 'Date', 'BU'], axis=1, inplace=True)




    # copy the catalog dataframe to a new dataframe called order_planning
    disway_catalog = catalog.copy()
    # create a new column called 'distriutor' and set it to 'Disway'
    disway_catalog['Distributor'] = 'DISWAY'

    disty_catalog = catalog.copy()
    # create a new column called 'distriutor' and set it to 'Disty'
    disty_catalog['Distributor'] = 'DISTY'

    # concatenate disty_catalog with disway_catalog
    order_planning = pd.concat([disty_catalog, disway_catalog])

    # # create a new column called '13 week sales' and set it to nan
    # order_planning['13 week sales'] = 0

    # # create a new column called 'delevery time' and set it to nan
    # order_planning['delevery time'] = 0

    # # create a new column called 'inventory' and set it to nan
    # order_planning['inventory'] = 0


    # merge order_planning with average_data on Part number and distributor columns
    order_planning = pd.merge(order_planning, average_data, how='left', left_on=['PartID', 'Distributor'], right_on=['PartID', 'Distributor'])

    # merge order_planning with backlog on Part number 
    order_planning = pd.merge(order_planning, backlog, how='left', left_on=['PartID'], right_on=['PartID'])

    # merge order_planning with inventory on Part number and distributor columns
    order_planning = pd.merge(order_planning, inventory, how='left', left_on=['PartID', 'Distributor'], right_on=['PartID', 'Distributor'])

    # rearanage the columns
    order_planning = order_planning[['PartID', 'Distributor', 'Inventory( Units)', 'Sold Qty', 'time_to_ship']]




    # for rows in order_planning dataframe where the inventory is nan but  Sold Qty is not nan then set the inventory to 0
    order_planning.loc[(order_planning['Inventory( Units)'].isna()) & (order_planning['Sold Qty'].notna()), 'Inventory( Units)'] = 0

    # for rows in order_planning dataframe where the Sold Qty is nan but  inventory is not nan then set the Sold Qty to 0
    order_planning.loc[(order_planning['Sold Qty'].isna()) & (order_planning['Inventory( Units)'].notna()), 'Sold Qty'] = 0



    # create a row called weekly sellout equal to Sold Qty / 13
    order_planning['weekly sellout'] = order_planning['Sold Qty'] / 13

    # create a row called weeks of sales equal to inventory / weekly sellout
    order_planning['weeks of sales'] = order_planning['Inventory( Units)'] / order_planning['weekly sellout']

    # create a row called shortage point equal to todays date + weeks of sales
    order_planning['shortage point'] = pd.to_datetime('today') + pd.to_timedelta(order_planning['weeks of sales'], unit='w')

    # create a row called reorder point equal to shortage point - time to ship(weeks)
    order_planning['reorder point'] = order_planning['shortage point'] - pd.to_timedelta(order_planning['time_to_ship'], unit='w')

    # create a row called status equal to 'FALSE' if reorder point is greater than todays date, 'TURE' if reorder point is less than todays date and 'UNKNOWN' if reorder point is nan
    order_planning['status'] = np.where(order_planning['reorder point'] > pd.to_datetime('today'), 'OK', np.where(order_planning['reorder point'] < pd.to_datetime('today'), 'ORDER NOW', 'UNKNOWN'))

    # if weeks of sales is not nan but time to ship is nan then set status to 'NO RECORDS IN BACKLOG'
    order_planning.loc[(order_planning['weeks of sales'].notna()) & (order_planning['time_to_ship'].isna()), 'status'] = 'NO RECORDS IN BACKLOG'

    # if weeks of sales is inf then set status to 'DOES NOT SELL'
    order_planning.loc[order_planning['weeks of sales'] == np.inf, 'status'] = 'DOES NOT SELL'

    # create a row called weeks left to order equal to weeks of sales - time to ship
    order_planning['weeks left to order'] = order_planning['weeks of sales'] - order_planning['time_to_ship']

    # add the product description to the order_planning dataframe from the backlogs dataframe based on the partID
    order_planning['Product Description'] = order_planning['PartID'].map(backlog.set_index('PartID')['Product Description'])



    # # add a column where the value is the number of units that has been ordered in the backlog_backups where the 'item status' column is either 'Processing' or 'In Transit' or 'Production' summed by partID if the 'Ship To Name' contains the distributor name
    # order_planning['units ordered'] = order_planning['PartID'].map(backlog_backups[backlog_backups['Ship To Name'].str.contains(order_planning['Distributor']) & backlog_backups['Item Status'].isin(['Processing', 'In Transit', 'Production'])].groupby('PartID')['Quantity'].sum())

    # do the same for every status and capitalise the ship to name
    order_planning['units processing'] = order_planning.apply(lambda row: backlog_backups[(backlog_backups['PartID'] == row['PartID']) & (backlog_backups['Ship To Name'].str.title().str.contains(row['Distributor'].title())) & (backlog_backups['Item Status'].isin(['Processing']))]['Ordered Quantity'].sum(), axis=1)
    order_planning['units in transit'] = order_planning.apply(lambda row: backlog_backups[(backlog_backups['PartID'] == row['PartID']) & (backlog_backups['Ship To Name'].str.title().str.contains(row['Distributor'].title())) & (backlog_backups['Item Status'].isin(['In Transit']))]['Ordered Quantity'].sum(), axis=1)
    order_planning['units production'] = order_planning.apply(lambda row: backlog_backups[(backlog_backups['PartID'] == row['PartID']) & (backlog_backups['Ship To Name'].str.title().str.contains(row['Distributor'].title())) & (backlog_backups['Item Status'].isin(['Production']))]['Ordered Quantity'].sum(), axis=1)

    # units ordered is the sum of units production units in transit and units processing
    order_planning['units ordered'] = order_planning['units production'] + order_planning['units in transit'] + order_planning['units processing']

    # if units ordered is not 0 and the status is 'ORDER NOW' then set status to 'ONGOING ORDER'
    order_planning.loc[(order_planning['units ordered'] != 0) & (order_planning['status'] == 'ORDER NOW'), 'status'] = 'ONGOING ORDER'

    # if status is OK and units ordered is 0 then set status to 'SAFE'
    order_planning.loc[(order_planning['status'] == 'OK') & (order_planning['units ordered'] == 0), 'status'] = 'SAFE'

    # create a column called priority equal to 4 if status is 'OK', 0 if status is 'ORDER NOW', 1 if status is 'ONGOING ORDER', 3 if status is 'NO RECORDS IN BACKLOG', 2 if status is 'DOES NOT SELL', 5 if status is 'UNKNOWN'
    order_planning['priority'] = np.where(order_planning['status'] == 'ORDER NOW', 0, np.where(order_planning['status'] == 'ONGOING ORDER', 1, np.where(order_planning['status'] == 'NO RECORDS IN BACKLOG', 3, np.where(order_planning['status'] == 'DOES NOT SELL', 2, np.where(order_planning['status'] == 'UNKNOWN', 5, 4)))))




    # reorder the columns
    order_planning = order_planning[['PartID', 'Product Description', 'Distributor', 'Inventory( Units)', 'Sold Qty', 'weekly sellout', 'weeks of sales', 'time_to_ship', 'shortage point', 'reorder point', 'weeks left to order', 'units ordered', 'status', 'units processing', 'units in transit', 'units production', 'priority']]

    # fill the product description column from backlog_backups if its nan based on the partID grouped by partID 
    order_planning['Product Description'] = order_planning['Product Description'].fillna(order_planning['PartID'].map(backlog_backups.groupby('PartID')['Product Description'].first()))
    # sort by status
    order_planning = order_planning.sort_values(by=['priority'])

    # display the order_planning dataframe
    st.write(order_planning)

    import io
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    import base64

    def get_formatted_excel(df):
        # Create a memory buffer
        buffer = io.BytesIO()
        
        # Convert DataFrame to Excel file
        df.to_excel(buffer, index=False)
        buffer.seek(0)
        # Load the Excel file
        wb = openpyxl.load_workbook(buffer)
        ws = wb.active
        
        # Iterate through the rows of the sheet
        for row in ws.iter_rows():
            # Check the value of the status column
            if row[12].value == 'ONGOING ORDER':
                # If the value is 'ONGOING ORDER', set the fill of the cell to blue
                row[12].fill = PatternFill(start_color="5fa5de", end_color="5fa5de", fill_type = "solid")
            elif row[12].value == 'OK':
                # If the value is 'OK', set the fill of the cell to green
                row[12].fill = PatternFill(start_color="6cde5f", end_color="6cde5f", fill_type = "solid")
            elif row[12].value == 'DOES NOT SELL':
                # If the value is 'DOES NOT SELL', set the fill of the cell to yellow
                row[12].fill = PatternFill(start_color="dec75f", end_color="dec75f", fill_type = "solid")
            elif row[12].value == 'ORDER NOW':
                # If the value is 'ORDER NOW', set the fill of the cell to red
                row[12].fill = PatternFill(start_color="de6c5f", end_color="de6c5f", fill_type = "solid")
            elif row[12].value == 'SAFE':
                # If the value is 'ORDER NOW', set the fill of the cell to red
                row[12].fill = PatternFill(start_color="5fdebc", end_color="5fdebc", fill_type = "solid")
        columns = [ws['E'], ws['F'], ws['G'], ws['H'], ws['D'], ws['L'], ws['K'], ws['O'], ws['P'], ws['N']]
        for column in columns:
            # Iterate through the cells in the column
            for cell in column:
                # Set the number format of the cell to comma style
                cell.number_format = '#,##0.00'
        columns2 = [ws['J'], ws['I']]
        for column in columns2:
            # Iterate through the cells in the column
            for cell in column:
                # Set the number format of the cell to comma style
                cell.number_format = 'dd/mm/yyyy'
        # delete the Q column
        ws.delete_cols(17)
        # add filter
        ws.auto_filter.ref = "A1:P1"
        # Seek the beginning of the buffer
        buffer.seek(0)
        # Save the Excel file to the buffer
        wb.save(buffer)
        # Encode the buffer as a base64 string
        encoded_data = base64.b64encode(buffer.getvalue()).decode()
        # Create the data URI
        data_uri = f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{encoded_data}'
        # Create a link with the data URI as the href
        st.markdown(f'<a href="{data_uri}" download="order_planning.xlsx">Download Excel</a>', unsafe_allow_html=True)

    get_formatted_excel(order_planning)
