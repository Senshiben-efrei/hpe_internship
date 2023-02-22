import streamlit as st
import pandas as pd
import time
import numpy as np

import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
import matplotlib.pyplot as plt

import io
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import base64

import requests
ip = '127.0.0.1:5000'
# create a function to get the data
def get_data():
    url = 'http://' + ip + '/api/v1/resources/data/all'
    response = requests.get(url)
    return response.json()

# create a function to add the data
def add_data(data):
    url = 'http://' + ip + '/api/v1/resources/data/add'
    response = requests.post(url, json=data)
    return response.json()

# create a function to delete the data
def delete_data(index):
    url = 'http://' + ip + '/api/v1/resources/data/delete/' + str(index)
    response = requests.delete(url)
    return response.json()

# create a function to update the data
def update_data(index, data):
    url = 'http://' + ip + '/api/v1/resources/data/update/' + str(index)
    response = requests.put(url, json=data)
    return response.json()

def create_dict(row):
    return {
        'purchase order number': row['purchase order number'],
        'Partner': row['Partner'],
        'Distributor': row['Distributor'],
        'Client': row['Client'],
        'Bundle config id' : row['Bundle config id'],
        'Product number': row['Product'],
        'quantity': row['Quantity'],
        'Description': row['Description'],
        'Unit Price': row['Price'],
        'Total Cost': row['Total']
    }




PAGE_DICT = {
    "Data Upload": "Data Upload",
    "Inventory status": "Inventory status",
    "Overview": "Data Overview",
    "Delivery time": "Delivery time",
    "Database": "Database",
}

DESCRIPTION_DICT = {
    "Data Upload": "💡 This tool is used to upload the data needed for all the other tools",
    "Overview": "💡 This tool will give you an insight on the data uploaded",
    "Inventory status": "💡 This tool calculates the sellout and weeks of sales and provides an order status for each product on the catalog",
    "Delivery time": "💡 This tool calculates the delivery time, orders and sold units for the inputted product number",
    "Database": "💡 This tool allows you to view, add and edit the data in the local database",
}
    
    
# # create empty session state variables
# st.session_state.catalog = None
# st.session_state.backlog = None
# st.session_state.sellout = None
# st.session_state.inventory = None
# st.session_state.backlog_backup = None


def download_database():
    # get the data from the api
    database_df = pd.DataFrame(get_data())
    # create a excel file
    excel_file = io.BytesIO()
    # write the dataframe to the excel file
    database_df.to_excel(excel_file, index=False)
    # create a download link
    excel_file.seek(0)
    b64 = base64.b64encode(excel_file.read()).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="database.xlsx">Download database</a>'
    st.markdown(href, unsafe_allow_html=True)

def get_formatted_excel(df):
    

    # Create a memory buffer
    buffer = io.BytesIO()
    
    # Convert DataFrame to Excel file
    df.to_excel(buffer, index=False)
    buffer.seek(0)
    # Load the Excel file
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active
    
    # Iterate through the rows of the sheet
    for row in ws.iter_rows():
        # Check the value of the status column
        if row[12].value == 'ONGOING ORDER':
            # If the value is 'ONGOING ORDER', set the fill of the cell to blue
            row[12].fill = PatternFill(start_color="5fa5de", end_color="5fa5de", fill_type = "solid")
        elif row[12].value == 'OK':
            # If the value is 'OK', set the fill of the cell to green
            row[12].fill = PatternFill(start_color="6cde5f", end_color="6cde5f", fill_type = "solid")
        elif row[12].value == 'DOES NOT SELL':
            # If the value is 'DOES NOT SELL', set the fill of the cell to yellow
            row[12].fill = PatternFill(start_color="dec75f", end_color="dec75f", fill_type = "solid")
        elif row[12].value == 'ORDER NOW':
            # If the value is 'ORDER NOW', set the fill of the cell to red
            row[12].fill = PatternFill(start_color="de6c5f", end_color="de6c5f", fill_type = "solid")
        elif row[12].value == 'SAFE':
            # If the value is 'ORDER NOW', set the fill of the cell to red
            row[12].fill = PatternFill(start_color="5fdebc", end_color="5fdebc", fill_type = "solid")
    columns = [ws['E'], ws['F'], ws['G'], ws['H'], ws['D'], ws['L'], ws['K'], ws['O'], ws['P'], ws['N']]
    for column in columns:
        # Iterate through the cells in the column
        for cell in column:
            # Set the number format of the cell to comma style
            cell.number_format = '#,##0.00'
    columns2 = [ws['J'], ws['I']]
    for column in columns2:
        # Iterate through the cells in the column
        for cell in column:
            # Set the number format of the cell to comma style
            cell.number_format = 'dd/mm/yyyy'
    # delete the Q column
    ws.delete_cols(17)
    # add filter
    ws.auto_filter.ref = "A1:P1"
    # Seek the beginning of the buffer
    buffer.seek(0)
    # Save the Excel file to the buffer
    wb.save(buffer)
    # Encode the buffer as a base64 string
    encoded_data = base64.b64encode(buffer.getvalue()).decode()
    # Create the data URI
    data_uri = f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{encoded_data}'
    # Create a link with the data URI as the href
    st.markdown(f'<a href="{data_uri}" download="order_planning.xlsx">Download Excel</a>', unsafe_allow_html=True)


def color_order_status(s):
    color = '#5fa5de ' if s == 'ONGOING ORDER' else '#6cde5f' if s == 'OK' else '#dec75f' if s == 'DOES NOT SELL' else '#de6c5f' if s == 'ORDER NOW' else '#5fdebc'
    return f'background-color: {color}'

def data_upload():
    # create 4 columns
    col1, col2, col3, col4 = st.columns(4)

    # second file
    sellout_file = col1.file_uploader("Sellout", type=["xlsx", "xls", "xlsm", "xlsb"], key="sellout_file")
    if 'sellout' in st.session_state:
        col1.success("Sellout file uploaded")
    if sellout_file is not None and 'sellout' not in st.session_state:
        st.session_state.sellout = pd.read_excel(sellout_file)
        col1.success("Sellout file uploaded")

    # third file
    inventory_file = col2.file_uploader("Inventory", type=["xlsx", "xls", "xlsm", "xlsb"], key="inventory_file")
    if 'inventory' in st.session_state:
        col2.success("Inventory file uploaded")
    if inventory_file is not None and 'inventory' not in st.session_state:
        st.session_state.inventory = pd.read_excel(inventory_file)
        col2.success("Inventory file uploaded")

    # fourth file
    catalog_file = col3.file_uploader("Catalog", type=["xlsx", "xls", "xlsm", "xlsb"], key="catalog_file")
    if 'catalog' in st.session_state:
        col3.success("Catalog file uploaded")
    if catalog_file is not None and 'catalog' not in st.session_state:
        st.session_state.catalog = pd.read_excel(catalog_file)
        col3.success("Catalog file uploaded")
    
    # disable file uploader if session state is not None
    backlog_backup_file = col4.file_uploader("Backlog", type=["xlsx", "xls", "xlsm", "xlsb"], key="backlog_backup_file")
    if 'backlog_backup' in st.session_state:
        col4.success("Backlog file uploaded")
    if backlog_backup_file is not None and 'backlog_backup' not in st.session_state:
        st.session_state.backlog_backup = pd.read_excel(backlog_backup_file)
        # copy the backlog_backup to a new dataframe
        backlog = st.session_state.backlog_backup.copy()
        # store the backlog in the session state
        st.session_state.backlog = backlog
        col4.success("Backlog file uploaded")

    # select country from all 'ma', 'ke'
    country = st.selectbox("Select country", ['Morroco', 'Kenya', 'Nigeria', 'Bulgaria', 'United Arab Emirates'])
    if country == 'Morroco':
        st.session_state.country = 'MA'
    elif country == 'Kenya':
        st.session_state.country = 'KE'
    elif country == 'Nigeria':
        st.session_state.country = 'NG'
    elif country == 'Bulgaria':
        st.session_state.country = 'BG'
    elif country == 'United Arab Emirates':
        st.session_state.country = 'AE'
    else:
        st.session_state.country = 'MA'


    # create a spacer
    st.markdown("---")

    # create a message to show if all the files are uploaded
    if 'catalog' in st.session_state and 'sellout' in st.session_state and 'inventory' in st.session_state and 'backlog' in st.session_state:
        st.subheader("All the files are uploaded. You can now use all the tools")

def overview():

    
    # if all the files are uploaded
    if 'backlog' in st.session_state:
        # create a copy of the dataframes
        backlog = st.session_state.backlog.copy()
        backlog_backup = st.session_state.backlog_backup.copy()
        
        if 'inventory' in st.session_state:
            inventory = st.session_state.inventory.copy()
        if 'sellout' in st.session_state:
            sellout = st.session_state.sellout.copy()
        if 'catalog' in st.session_state:
            catalog = st.session_state.catalog.copy()

            
        # drop Serial Number and Box ID rows
        backlog = backlog.drop(['Serial Number', 'Box ID'], axis=1)
        # drop duplicates rows
        backlog = backlog.drop_duplicates()

        
        # change the date columns to datetime Actual Delivery Date (Item) and HPE Received Date
        backlog['Actual Delivery Date (Item)'] = pd.to_datetime(backlog['Actual Delivery Date (Item)'])
        backlog['HPE Received Date'] = pd.to_datetime(backlog['HPE Received Date'])

        # actual delivery date is different from nan and 'N/A' change item status to 'Delivered'
        backlog.loc[(backlog['Actual Delivery Date (Item)'] != 'N/A') & (backlog['Actual Delivery Date (Item)'].notna()), 'Item Status'] = 'Delivered'
        

        # calculate the the time between the HPE Received Date and the Actual Delivery Date (Item)
        backlog['time_to_ship'] = backlog['Actual Delivery Date (Item)'] - backlog['HPE Received Date'] 

        # convert the time to days
        backlog['time_to_ship'] = backlog['time_to_ship'].dt.days/5

        # drop rows where time_to_ship is inf or nan
        backlog = backlog[backlog['time_to_ship'] != np.inf]
        backlog = backlog[backlog['time_to_ship'].notna()]

        # create a new column called 'shiped to' and assign the value of the column 'Ship To Name'
        backlog['shiped_to'] = backlog['Ship To Name']

        # if of shiped_to contains 'disway' or 'disty' then change the value to it
        backlog.loc[backlog['shiped_to'].str.contains('disway', case=False), 'shiped_to'] = 'disway'
        backlog.loc[backlog['shiped_to'].str.contains('disty', case=False), 'shiped_to'] = 'disty'

        # count the frequency of each value in the column 'Product Number' and plot the top 10
        st.subheader("Top 10 products by backlog")
        st.bar_chart(backlog['Product Number'].value_counts().head(10))

        # write the top 10 products in a table
        st.subheader("Top 10 products by backlog")
        st.table(backlog['Product Number'].value_counts().head(10))
        



                

    else:
        st.write("Please upload the backlog")
        st.write("👈 Go to the Data Upload tool to do so")

        # write in the sidebar the instructions alert message to upload the files
        st.sidebar.error("👆 You need at least the backlog file to use this tool")



def inventory_status():

    # check if there is a file named order_planning in the session state
    if 'order_planning' in st.session_state:
        # if there is a file, load it
        order_planning = st.session_state.order_planning
        # display the file
        st.write(order_planning)
        get_formatted_excel(order_planning)

    else:

            # if all the files are uploaded
        if 'catalog' in st.session_state and 'sellout' in st.session_state and 'inventory' in st.session_state and 'backlog' in st.session_state:
            
            # create a button to start the process
            if st.button("Start the process"):
                # create a progress bar
                progress_bar = st.progress(0)
                # create a status text
                status_text = st.empty()


                # create copies of the dataframes
                catalog = st.session_state.catalog
                backlog = st.session_state.backlog
                sellout = st.session_state.sellout
                inventory = st.session_state.inventory
                backlog_backup = st.session_state.backlog_backup

                # progress bar
                progress_bar.progress(10)
                status_text.text("loading files")

                #cleaning
                # change date format to datetime
                sellout['Date'] = pd.to_datetime(sellout['Date'])

                # progress bar
                progress_bar.progress(15)
                status_text.text("changing date format")

                # if the part number contains ' 0D1' or ' OD1' then replace it with ''
                sellout['Part number'] = sellout['Part number'].str.replace(' 0D1', '')
                sellout['Part number'] = sellout['Part number'].str.replace(' OD1', '')
                # in the distributor column, uppercase all the letters
                sellout['Distributor'] = sellout['Distributor'].str.upper()
                # remove the leading and trailing spaces from the distributor column
                sellout['Distributor'] = sellout['Distributor'].str.strip()
                # remove the leading and trailing spaces from the part number column
                sellout['Part number'] = sellout['Part number'].str.strip()

                # Set the index of the dataframe to the 'Date' column
                sellout.set_index('Date', inplace=True)

                # Group the data by 'Part number' and 'Distributor'
                grouped_data = sellout.groupby(['Part number', 'Distributor'])

                # progress bar
                progress_bar.progress(20)
                status_text.text("cleaning data")

                # Resample the data by week
                resampled_data = grouped_data.resample('W')

                # Aggregate the data by summing the 'Sold Qty' column
                aggregated_data = resampled_data['Sold Qty'].sum()

                # look for the last 13 weeks from today and average the data for each part number and distributor
                # Get todays date and subtract 13 weeks
                last_13_weeks = pd.to_datetime('today') - pd.DateOffset(weeks=13)

                # remove the data whose date is less than 13 weeks from today
                filtered_data = aggregated_data[aggregated_data.index.get_level_values('Date') > last_13_weeks]

                # average the data for each part number and distributor
                average_data = filtered_data.groupby(['Part number', 'Distributor']).sum()

                # Reset the index of the dataframe
                average_data = average_data.reset_index()

                # store the average data in a new dataframe with the columns 'Part number', 'Distributor', and 'Average'
                average_data = average_data[['Part number', 'Distributor', 'Sold Qty']]
                # rename the 'Part number' column to 'PartID'
                average_data.rename(columns={'Part number': 'PartID'}, inplace=True)



                # in the catalog dataframe drop the Row Labels column
                catalog.drop('Row Labels', axis=1, inplace=True)

                # progress bar
                progress_bar.progress(25)
                status_text.text("dropping duplicate rows")

                # drop duplicate rows
                catalog.drop_duplicates(inplace=True)

                # progress bar
                progress_bar.progress(30)
                status_text.text("calculating shipping time")


                # process the backlog data
                backlog['Actual Delivery Date (Item)'] = pd.to_datetime(backlog['Actual Delivery Date (Item)'])
                backlog['HPE Received Date'] = pd.to_datetime(backlog['HPE Received Date'])

                # calculate the the time between the HPE Received Date and the Actual Delivery Date (Item)
                backlog['time_to_ship'] = backlog['Actual Delivery Date (Item)'] - backlog['HPE Received Date'] 

                # convert the time to days
                backlog['time_to_ship'] = backlog['time_to_ship'].dt.days/5

                # drop rows where time_to_ship is inf or nan
                backlog = backlog[backlog['time_to_ship'] != np.inf]
                backlog = backlog[backlog['time_to_ship'].notna()]

                # drop rows where Order Type is not 'HPE Non Stkng Order'
                backlog = backlog[backlog['Order Type'] == 'HPE Stocking Order']

                # drop all comlumns except for 'Part number', 'shiped_to', 'time_to_ship', 'Material Description'
                backlog = backlog[['Product Number', 'time_to_ship', 'Product Description']]

                # calculate the average time to ship for each part number keeping the product description
                backlog = backlog.groupby(['Product Number', 'Product Description']).median()

                # reset the index keep product description
                backlog = backlog.reset_index()

                # drop all rows where product number is not in the catalog
                backlog = backlog[backlog['Product Number'].isin(catalog['PartID'])]

                # rename the 'Product Number' column to 'PartID'
                backlog.rename(columns={'Product Number': 'PartID'}, inplace=True)
                backlog_backup.rename(columns={'Product Number': 'PartID'}, inplace=True)

                # in backlog_backup, change Item Status to 'Delivered' if the Actual Delivery Date (Item) is not null
                backlog_backup['Item Status'] = np.where(backlog_backup['Actual Delivery Date (Item)'].notna(), 'Delivered', backlog_backup['Item Status'])



                # in inventory dataframe keep this columns  Week	Month	Date	Distributor	BU	Part number	Inventory( Units)
                inventory = inventory[['Week', 'Month', 'Date', 'Distributor', 'BU', 'Part number', 'Inventory( Units)']]

                # if the part number contains ' 0D1' or ' OD1' then replace it with ''
                inventory['Part number'] = inventory['Part number'].str.replace(' 0D1', '')
                inventory['Part number'] = inventory['Part number'].str.replace(' OD1', '')
                # in the distributor column, uppercase all the letters
                inventory['Distributor'] = inventory['Distributor'].str.upper()
                # remove the leading and trailing spaces from the distributor column
                inventory['Distributor'] = inventory['Distributor'].str.strip()
                # remove the leading and trailing spaces from the part number column
                inventory['Part number'] = inventory['Part number'].str.strip()

                # progress bar
                progress_bar.progress(35)
                status_text.text("processing inventory data")
                

                # remove the first letter in week column 
                inventory['Week'] = inventory['Week'].str[1:]

                # convert the week column to integer
                inventory['Week'] = inventory['Week'].astype(int)

                # look for the latest Week in the inventory dataframe
                latest_week = inventory['Week'].max()

                # filter the inventory dataframe to keep only the latest week
                inventory = inventory[inventory['Week'] == latest_week]

                # limit the inventory dataframe to catalog part numbers
                inventory = inventory[inventory['Part number'].isin(catalog['PartID'])]

                # rename the column part number to PartID
                inventory.rename(columns={'Part number': 'PartID'}, inplace=True)

                # drop Week	Month	Date		BU
                inventory.drop(['Week', 'Month', 'Date', 'BU'], axis=1, inplace=True)

                # progress bar
                progress_bar.progress(40)
                status_text.text("processing catalog data")




                # copy the catalog dataframe to a new dataframe called order_planning
                disway_catalog = catalog.copy()
                # create a new column called 'distriutor' and set it to 'Disway'
                disway_catalog['Distributor'] = 'DISWAY'

                disty_catalog = catalog.copy()
                # create a new column called 'distriutor' and set it to 'Disty'
                disty_catalog['Distributor'] = 'DISTY'

                # concatenate disty_catalog with disway_catalog
                order_planning = pd.concat([disty_catalog, disway_catalog])

                # # create a new column called '13 week sales' and set it to nan
                # order_planning['13 week sales'] = 0

                # # create a new column called 'delevery time' and set it to nan
                # order_planning['delevery time'] = 0

                # # create a new column called 'inventory' and set it to nan
                # order_planning['inventory'] = 0

                # progress bar
                progress_bar.progress(45)
                status_text.text("creating order planning dataframe")


                # merge order_planning with average_data on Part number and distributor columns
                order_planning = pd.merge(order_planning, average_data, how='left', left_on=['PartID', 'Distributor'], right_on=['PartID', 'Distributor'])

                # merge order_planning with backlog on Part number 
                order_planning = pd.merge(order_planning, backlog, how='left', left_on=['PartID'], right_on=['PartID'])

                # merge order_planning with inventory on Part number and distributor columns
                order_planning = pd.merge(order_planning, inventory, how='left', left_on=['PartID', 'Distributor'], right_on=['PartID', 'Distributor'])

                # rearanage the columns
                order_planning = order_planning[['PartID', 'Distributor', 'Inventory( Units)', 'Sold Qty', 'time_to_ship']]

                # progress bar
                progress_bar.progress(50)
                status_text.text("processing order planning dataframe")


                # for rows in order_planning dataframe where the inventory is nan but  Sold Qty is not nan then set the inventory to 0
                order_planning.loc[(order_planning['Inventory( Units)'].isna()) & (order_planning['Sold Qty'].notna()), 'Inventory( Units)'] = 0

                # for rows in order_planning dataframe where the Sold Qty is nan but  inventory is not nan then set the Sold Qty to 0
                order_planning.loc[(order_planning['Sold Qty'].isna()) & (order_planning['Inventory( Units)'].notna()), 'Sold Qty'] = 0

                # progress bar
                progress_bar.progress(55)
                status_text.text("calculating sellout ")


                # create a row called weekly sellout equal to Sold Qty / 13
                order_planning['weekly sellout'] = order_planning['Sold Qty'] / 13

                # create a row called weeks of sales equal to inventory / weekly sellout
                order_planning['weeks of sales'] = order_planning['Inventory( Units)'] / order_planning['weekly sellout']

                # create a row called shortage point equal to todays date + weeks of sales
                order_planning['shortage point'] = pd.to_datetime('today') + pd.to_timedelta(order_planning['weeks of sales'], unit='w')

                # create a row called reorder point equal to shortage point - time to ship(weeks)
                order_planning['reorder point'] = order_planning['shortage point'] - pd.to_timedelta(order_planning['time_to_ship'], unit='w')

                # create a row called status equal to 'FALSE' if reorder point is greater than todays date, 'TURE' if reorder point is less than todays date and 'UNKNOWN' if reorder point is nan
                order_planning['status'] = np.where(order_planning['reorder point'] > pd.to_datetime('today'), 'OK', np.where(order_planning['reorder point'] < pd.to_datetime('today'), 'ORDER NOW', 'UNKNOWN'))

                # if weeks of sales is not nan but time to ship is nan then set status to 'NO RECORDS IN BACKLOG'
                order_planning.loc[(order_planning['weeks of sales'].notna()) & (order_planning['time_to_ship'].isna()), 'status'] = 'NO RECORDS IN BACKLOG'

                # if weeks of sales is inf then set status to 'DOES NOT SELL'
                order_planning.loc[order_planning['weeks of sales'] == np.inf, 'status'] = 'DOES NOT SELL'

                # create a row called weeks left to order equal to weeks of sales - time to ship
                order_planning['weeks left to order'] = order_planning['weeks of sales'] - order_planning['time_to_ship']

                # add the product description to the order_planning dataframe from the backlogs dataframe based on the partID
                order_planning['Product Description'] = order_planning['PartID'].map(backlog.set_index('PartID')['Product Description'])

                # progress bar
                progress_bar.progress(60)
                status_text.text("looking for processing orders in the backlog")



                # # add a column where the value is the number of units that has been ordered in the backlog_backup where the 'item status' column is either 'Processing' or 'In Transit' or 'Production' summed by partID if the 'Ship To Name' contains the distributor name
                # order_planning['units ordered'] = order_planning['PartID'].map(backlog_backup[backlog_backup['Ship To Name'].str.contains(order_planning['Distributor']) & backlog_backup['Item Status'].isin(['Processing', 'In Transit', 'Production'])].groupby('PartID')['Quantity'].sum())

                # do the same for every status and capitalise the ship to name
                order_planning['units processing'] = order_planning.apply(lambda row: backlog_backup[(backlog_backup['PartID'] == row['PartID']) & (backlog_backup['Ship To Name'].str.title().str.contains(row['Distributor'].title())) & (backlog_backup['Item Status'].isin(['Processing']))]['Ordered Quantity'].sum(), axis=1)
                # progress bar
                progress_bar.progress(65)
                status_text.text("looking for transit orders in the backlog")

                order_planning['units in transit'] = order_planning.apply(lambda row: backlog_backup[(backlog_backup['PartID'] == row['PartID']) & (backlog_backup['Ship To Name'].str.title().str.contains(row['Distributor'].title())) & (backlog_backup['Item Status'].isin(['In Transit']))]['Ordered Quantity'].sum(), axis=1)
                # progress bar
                progress_bar.progress(70)
                status_text.text("looking for production orders in the backlog")

                order_planning['units production'] = order_planning.apply(lambda row: backlog_backup[(backlog_backup['PartID'] == row['PartID']) & (backlog_backup['Ship To Name'].str.title().str.contains(row['Distributor'].title())) & (backlog_backup['Item Status'].isin(['Production']))]['Ordered Quantity'].sum(), axis=1)

                # units ordered is the sum of units production units in transit and units processing
                order_planning['units ordered'] = order_planning['units production'] + order_planning['units in transit'] + order_planning['units processing']

                # if units ordered is not 0 and the status is 'ORDER NOW' then set status to 'ONGOING ORDER'
                order_planning.loc[(order_planning['units ordered'] != 0) & (order_planning['status'] == 'ORDER NOW'), 'status'] = 'ONGOING ORDER'

                # if status is OK and units ordered is 0 then set status to 'SAFE'
                order_planning.loc[(order_planning['status'] == 'OK') & (order_planning['units ordered'] == 0), 'status'] = 'SAFE'

                # create a column called priority equal to 4 if status is 'OK', 0 if status is 'ORDER NOW', 1 if status is 'ONGOING ORDER', 3 if status is 'NO RECORDS IN BACKLOG', 2 if status is 'DOES NOT SELL', 5 if status is 'UNKNOWN'
                order_planning['priority'] = np.where(order_planning['status'] == 'ORDER NOW', 0, np.where(order_planning['status'] == 'ONGOING ORDER', 1, np.where(order_planning['status'] == 'NO RECORDS IN BACKLOG', 3, np.where(order_planning['status'] == 'DOES NOT SELL', 2, np.where(order_planning['status'] == 'UNKNOWN', 5, 4)))))

                # progress bar
                progress_bar.progress(75)
                status_text.text("calculating reorder point")


                # reorder the columns
                order_planning = order_planning[['PartID', 'Product Description', 'Distributor', 'Inventory( Units)', 'Sold Qty', 'weekly sellout', 'weeks of sales', 'time_to_ship', 'shortage point', 'reorder point', 'weeks left to order', 'units ordered', 'status', 'units processing', 'units in transit', 'units production', 'priority']]

                # fill the product description column from backlog_backup if its nan based on the partID grouped by partID 
                order_planning['Product Description'] = order_planning['Product Description'].fillna(order_planning['PartID'].map(backlog_backup.groupby('PartID')['Product Description'].first()))
                # sort by status
                order_planning = order_planning.sort_values(by=['priority'])

                # progress bar
                progress_bar.progress(80)
                status_text.text("dataframe ready")

                # store the order_planning dataframe in a session state
                st.session_state.order_planning = order_planning

                # display the order_planning dataframe
                st.write(order_planning)

                # progress bar
                progress_bar.progress(90)
                status_text.text("exporting to excel")

                # progress bar
                progress_bar.progress(100)
                status_text.text("Done!")

                get_formatted_excel(order_planning)

        else:
            st.write("Please upload all the files")
            st.write("👈 Go to the Data Upload tool to do so")

            # write in the sidebar the instructions alert message to upload the files
            st.sidebar.error("👆 You need all the 4 files to use this tool")

def delivery_time():
    # if all the files are uploaded
    if 'backlog' in st.session_state:
        # create a copy of the dataframes
        backlog = st.session_state.backlog.copy()
        backlog_backup = st.session_state.backlog_backup.copy()
        
        if 'inventory' in st.session_state:
            inventory = st.session_state.inventory.copy()
        if 'sellout' in st.session_state:
            sellout = st.session_state.sellout.copy()
        if 'catalog' in st.session_state:
            catalog = st.session_state.catalog.copy()

        # drop Serial Number and Box ID rows
        backlog = backlog.drop(['Serial Number', 'Box ID'], axis=1)
        # drop duplicates rows
        backlog = backlog.drop_duplicates()

        # drop all rows where ship to country is not MA
        backlog = backlog[backlog['Ship To Country'] == st.session_state.country]


        # change the date columns to datetime Actual Delivery Date (Item) and HPE Received Date
        backlog['Actual Delivery Date (Item)'] = pd.to_datetime(backlog['Actual Delivery Date (Item)'])
        backlog['HPE Received Date'] = pd.to_datetime(backlog['HPE Received Date'])

        # if row have actual delivery date change item status to 'Delivered' and sumariezed status to 'Delivered'
        backlog.loc[backlog['Actual Delivery Date (Item)'].notna(), 'Summarized Status (header level)'] = 'Delivered'
        backlog.loc[backlog['Actual Delivery Date (Item)'].notna(), 'Item Status'] = 'Delivered'
        

        # calculate the the time between the HPE Received Date and the Actual Delivery Date (Item)
        backlog['time_to_ship'] = backlog['Actual Delivery Date (Item)'] - backlog['HPE Received Date'] 

        # convert the time to days
        backlog['time_to_ship'] = backlog['time_to_ship'].dt.days/7

        # drop rows where time_to_ship is inf or nan
        backlog = backlog[backlog['time_to_ship'] != np.inf]
        backlog = backlog[backlog['time_to_ship'].notna()]


        # text input for product description
        product_number = st.text_input('Input Product Number', placeholder='Example : P16926-421')
        # submit button
        submit = st.button('Submit')

        # if submit button is pressed
        if submit or product_number:
            # if the text input is is surronded by quotes remove them
            if product_number[0] == '"' and product_number[-1] == '"':
                product_number = product_number[1:-1]

            
            if 'catalog' in st.session_state:
                #if the product_number is not in the catalog write a warning
                if product_number not in catalog['PartID'].values:
                    st.warning('Product not in catalog')
                else:
                    st.success('Product available in catalog')
            else:
                st.sidebar.warning('➕ Tip : Upload the catalog file to check directly if the product is in it')

            # write the product description
            st.write('Product Description : ' + backlog[backlog['Product Number'] == product_number]['Product Description'].values[0])

            # # write the mean time to ship for product_number # rounded to 2 decimals
            # st.write('Average time to ship : ' + str(round(backlog[backlog['Product Number'] == product_number]['time_to_ship'].mean(), 2)) + ' full weeks (7 days)')

            # # write the mean time to ship for product_number # rounded to 2 decimals working weeks (5 days)
            # st.write('Average time to ship : ' + str(round(backlog[backlog['Product Number'] == product_number]['time_to_ship'].mean()*7/5, 2)) + ' working weeks (5 days)')

            # # write number of orders for product_number
            # st.write('Number of orders : ' + str(backlog[backlog['Product Number'] == product_number]['Product Number'].value_counts()[product_number]))

            # # write average ordered quantity by order for product_number # rounded to 2 decimals
            # st.write('Average ordered quantity by order : ' + str(round(backlog[backlog['Product Number'] == product_number]['Ordered Quantity'].mean(), 2)))

            # # write sold units for product_number by summing the sold units and thier quantity 'Ordered Quantity'
            # st.write('Sold units : ' + str(backlog[backlog['Product Number'] == product_number]['Ordered Quantity'].sum()))

            # form the previous writes to a table
            table = pd.DataFrame({'Average time to ship (full weeks)': [round(backlog[backlog['Product Number'] == product_number]['time_to_ship'].mean(), 2)],
                                'Average time to ship (working weeks)': [round(backlog[backlog['Product Number'] == product_number]['time_to_ship'].mean()*7/5, 2)],
                                'Number of orders': [backlog[backlog['Product Number'] == product_number]['Product Number'].value_counts()[product_number]],
                                'Average quantity by order': [round(backlog[backlog['Product Number'] == product_number]['Ordered Quantity'].mean(), 2)],
                                'Sold units': [backlog[backlog['Product Number'] == product_number]['Ordered Quantity'].sum()]})
            # write the table vertically with bigger font with 2 decimals
            st.table(table.T.style.set_properties(**{'font-size': '20px'}).format("{:.2f}"))

            # turn date into this format "10 oct 2020"
            backlog['Actual Delivery Date (Item)'] = backlog['Actual Delivery Date (Item)'].dt.strftime('%b %d %Y')

            
            # plot line plot of time to ship for product_number 
            fig = px.scatter(backlog[backlog['Product Number'] == product_number], x='HPE Received Date', y='time_to_ship', color='Product Description')
            # add sold units to the hover data
            fig.update_traces(hovertemplate='HPE Received Date: %{x}<br>Delivery Date: %{customdata[2]}<br>Time to Deliver(7 days weeks): %{y}<br>Ordered units: %{customdata[0]}<br>Sold to: %{customdata[1]}<extra></extra> ', customdata=backlog[backlog['Product Number'] == product_number][['Ordered Quantity', 'Ship To Name', 'Actual Delivery Date (Item)']].values)
            fig.update_layout(title='Orders: time to ship and quantity')
            # change dot color and hover color to #01A982
            fig.update_traces(marker_color='#01A982', hoverlabel_bgcolor='#01A982')
            # change background dot size to 5
            fig.update_traces(marker_size=10)
            st.plotly_chart(fig)

            # write a dataframe backlog with this product_number
            st.write('Backlog preview')
            st.write(backlog[backlog['Product Number'] == product_number])

            
            if 'sellout' in st.session_state:
                # write a dataframe sellout with this product_number
                st.write('Sellout preview')
                st.write(sellout[sellout['Part number'] == product_number])
            else:
                st.sidebar.warning('➕ Tip : Upload the sellout file to check the latest sold units')


            if 'inventory' in st.session_state:
                # write a dataframe inventory with this product_number
                st.write('Inventory preview')
                st.write(inventory[inventory['Part number'] == product_number])
            else:
                st.sidebar.warning('➕ Tip : Upload the inventory file to check distributor inventory')
                

    else:
        st.write("Please upload the backlog")
        st.write("👈 Go to the Data Upload tool to do so")

        # write in the sidebar the instructions alert message to upload the files
        st.sidebar.error("👆 You need at least the backlog file to use this tool")

def database():
    # get the data from the api
    database_df = pd.DataFrame(get_data())
    # write the dataframe
    st.write(database_df)
    # create a button to refresh the data
    if st.button('Refresh'):
        # write refresh message
        st.success('Refreshed data')

    # create 3 columns
    col1, col2, col3 = st.columns(3)

    # create a form to add new data
    with col1.form('Add new data'):
        # write the title
        st.success('Add new data')
        # fill the following data 'partner' 'distributor' 'client' 'Bundle config id'	'Product number'	'Component quantity'	'Description'	'Unit Cost Price (USD)'	'Total Cost Price (USD)'	'Unit Selling Price (USD)'	'Total Selling Price (USD)'
        # create a text input for the purchase order number
        purchase_order_number = st.text_input('purchase order number')
        # create a text input for the partner
        partner = st.text_input('Partner')
        # create a text input for the distributor
        distributor = st.text_input('Distributor')
        # create a text input for the client
        client = st.text_input('Client')
        # paragraph input
        input_text = st.text_area('Enter the text')

        # submit button
        submit_button = st.form_submit_button('Submit')

        # if the submit button is clicked
        if submit_button:
            # write a success message
            st.success('Added data')
            # refresh the page
            # st.experimental_rerun()
            # create a file-like object from the input text
            input_file = io.StringIO(input_text)

            # read the text into a pandas dataframe
            new_order_df = pd.read_csv(input_file, sep='\t', header=None, names=['Bundle config id', 'Product', 'Quantity', 'Description', 'Price', 'Total'])

            # add the inputed data to the dataframe of each row
            new_order_df['purchase order number'] = purchase_order_number
            new_order_df['Partner'] = partner
            new_order_df['Distributor'] = distributor
            new_order_df['Client'] = client

            new_order_df = new_order_df.fillna('')
            

            # display the dataframe in the Streamlit app
            st.dataframe(new_order_df)

            # list of dicrionaries to add to the database
            list_of_dicts = new_order_df.apply(create_dict, axis=1).tolist()

            st.write(list_of_dicts)

            # add the data to the database loop through the list of dictionaries
            for dict_ in list_of_dicts:
                # add the data to the database
                add_data(dict_)

    # if the database is not empty
    if len(database_df) > 0:
        # create a form to update data
        with col2.form('Update data'):
            # write the index of the row to update
            st.warning('Input the index of the row to update')
            # create a integer input for the index unless the database is empty
            index = st.number_input('Index', min_value=0, max_value=len(database_df)-1)
            # create a text input for the partner
            partner = st.text_input('Partner')
            # create a text input for the distributor
            distributor = st.text_input('Distributor')
            # create a text input for the client
            client = st.text_input('Client')
            # submit button
            submit_button = st.form_submit_button('Submit')

            # if the submit button is clicked
            if submit_button:
                # update the data in the database
                update_data(index, {'purchase order number' : partner, 'Partner' : partner, 'Distributor' : distributor, 'Client' : client, 'Bundle config id': bundle_config_id, 'Product number': product_number, 'Component quantity': component_quantity, 'Description': description, 'Unit Cost Price (USD)': unit_cost_price, 'Total Cost Price (USD)': total_cost_price, 'Unit Selling Price (USD)': unit_selling_price, 'Total Selling Price (USD)': total_selling_price, 'modified': 'last edit ' + str(pd.datetime.now().replace(microsecond=0))})
                # write a success message
                st.success('Updated data')
                # refresh the page
                st.experimental_rerun()

    # if the database is not empty
    if len(database_df) > 0:
        # create a form to delete data
        with col3.form('Delete data'):
            # write the index of the row to delete
            st.error('Input the index of the row to delete')
            # create a integer input for the index
            index = st.number_input('Index', min_value=0, max_value=len(database_df)-1)
            # submit button
            submit_button = st.form_submit_button('Submit')

            # if the submit button is clicked
            if submit_button:
                # delete the data from the database
                delete_data(index)
                # write a success message
                st.success('Data deleted')
                # refresh the page
                st.experimental_rerun()

    # create a button to download the database
    if st.button('Generate a download link'):
        # download the database
        download_database()



def main():
    st.set_page_config(page_title="My Streamlit App", page_icon=":guardsman:", layout="wide")
        
    selected_page = st.sidebar.selectbox("Choose a tool", list(PAGE_DICT.keys()))
    st.subheader(PAGE_DICT[selected_page])
    #write the description of the selected page in info box on sidebar
    st.sidebar.info(DESCRIPTION_DICT[selected_page])


    # create 3 pages
    if selected_page == "Data Upload":
        data_upload()
    elif selected_page == "Overview":
        overview()
    elif selected_page == "Inventory status":
        inventory_status()
    elif selected_page == "Delivery time":
        delivery_time()
    elif selected_page == "Database":
        database()
    
    # info : if you want to load new data refresh the page
    if 'backlog' in st.session_state or 'inventory' in st.session_state or 'catalog' in st.session_state or 'sellout' in st.session_state:
        st.info("In case you want to load new data or in case of app crash please refresh the page")



if __name__ == "__main__":
    main()
